import pandas as pd
import gc
import os
import zipfile
from openpyxl import load_workbook

def normalizar_ean(serie):
    return (
        serie.astype(str)                         # força string
            .astype(str)
            .str.extract(r'(\d+)')[0]   # pega só números
            .str[-13:]                  # mantém os 13 últimos dígitos
            .str.zfill(13)              # preenche zeros à esquerda se faltar
    )
def is_strict_openxml(filepath):
    try:
        with zipfile.ZipFile(filepath, 'r') as zipf:
            content_types = zipf.read('[Content_Types].xml').decode('utf-8')
            return 'Strict' in content_types
    except Exception:
        return False

def corrigir_strict_openxml(caminho_original, caminho_corrigido):
    try:
        wb = load_workbook(caminho_original)
        wb.save(caminho_corrigido)
        return True
    except Exception as e:
        print(f"Erro ao corrigir arquivo: {e}")
        return False
    
def tentar_ler_csv(filepath):
    encodings = ["utf-8", "latin1", "iso-8859-1", "cp1252"]
    separadores = [",", ";", "\t", "|"]
    for enc in encodings:
        for sep in separadores:
            try:
                df = pd.read_csv(filepath, encoding=enc, sep=sep)
                #print(f"CSV lido com encoding '{enc}' e separador '{sep}'")
                return df
            except Exception:
                continue
    raise ValueError(f"Falha ao ler CSV: {filepath} com encodings e separadores testados.")

def parse(file, caminho_completo, output_path, engine):
    if file.lower().endswith('.xlsx'):
        if is_strict_openxml(caminho_completo):
            print(f"Arquivo está no formato Strict Open XML: {file}")
            caminho_corrigido = caminho_completo.replace('.xlsx', '_corrigido.xlsx')
            if corrigir_strict_openxml(caminho_completo, caminho_corrigido):
                print(f"Corrigido e salvo como: {os.path.basename(caminho_corrigido)}")
                caminho_completo = caminho_corrigido
            else:
                print("Não foi possível corrigir automaticamente.")

        try:
            xls = pd.ExcelFile(caminho_completo, engine="openpyxl")
            for sheet in xls.sheet_names:
                print("Cliente - ARAUJO")
                print(f"Arquivo - {file} | Aba - {sheet}")
                
                if sheet.strip().lower() == "quantidade":
                    # Ler a aba ignorando as 9 primeiras linhas
                    df_raw = pd.read_excel(xls, sheet_name=sheet, header=None, skiprows=8)
                    
                    # Detectar número de colunas úteis
                    # Forçar uso de todas as colunas
                    n_cols_total = df_raw.shape[1]  # usa todas as colunas visíveis
                    # Cabeçalhos
                    header_ano = df_raw.iloc[1, :n_cols_total].ffill()
                    header_mes = df_raw.iloc[2, :n_cols_total].ffill()

                    # Dados
                    dados = df_raw.iloc[3:, :n_cols_total].reset_index(drop=True)
                    multi_index = pd.MultiIndex.from_arrays([header_ano, header_mes])
                    dados.columns = multi_index

                    # Processar os nomes de coluna corretamente
                    colunas_formatadas = []
                    colunas_vendas = []

                    for a, b in dados.columns:
                        try:
                            a_str = str(int(float(a))).zfill(4) if pd.notna(a) and str(a).replace('.', '', 1).isdigit() else str(a).strip()
                            b_str = str(int(float(b))).zfill(2) if pd.notna(b) and str(b).replace('.', '', 1).isdigit() else str(b).strip()
                        except Exception:
                            a_str = str(a).strip()
                            b_str = str(b).strip()

                        if a_str.isdigit() and b_str.isdigit():
                            nome_col = f"{a_str}_{b_str}"
                            colunas_vendas.append(nome_col)
                        else:
                            nome_col = b_str

                        colunas_formatadas.append(nome_col)

                    dados.columns = colunas_formatadas

                    # Realizar o melt
                    df_melted = dados.melt(
                        id_vars=[col for col in dados.columns if col not in colunas_vendas],
                        value_vars=colunas_vendas,
                        var_name='ANO_MES',
                        value_name='QTD VENDA'
                    )
                    # Verificação de segurança
                    if not df_melted.empty:
                        # Remover nulos e validar o formato do ANO_MES
                        df_melted = df_melted[df_melted['ANO_MES'].notna()]
                        df_melted = df_melted[df_melted['ANO_MES'].str.match(r'^\d{4}_\d{1,2}$')]

                        # Separar Ano e Mês
                        df_melted[['Ano', 'Mês']] = df_melted['ANO_MES'].str.split("_", expand=True)

                        # Criar campo DATA
                        df_melted['DATA'] = pd.to_datetime({
                            'year': df_melted['Ano'].astype(int),
                            'month': df_melted['Mês'].astype(int),
                            'day': 1
                        }).dt.strftime('%m/%d/%Y')
                        df_melted = df_melted[
                            df_melted['Produto'].notna() &
                            df_melted['Cód Barra Venda'].notna() &
                            df_melted['Situação Atual'].notna() &
                            df_melted['QTD VENDA'].notna()
                        ]
                    df2 = pd.DataFrame()
                    df2['MÊS'] = df_melted['DATA']
                    df2['CANAL'] = 'REDE'
                    df2['CLIENTE'] = 'ARAUJO'
                    df2['GC'] = 'ANDRE RODRIGUES'
                    df2['UF'] = 'MG'
                    df2['EAN'] = normalizar_ean(df_melted['Cód Barra Venda'].astype(str).str.zfill(13))
                    df2['FAMILIA'] = df_melted['Produto'].astype(str).str.strip().str.split().str[0]
                    df2['SKU_SERVIER'] = df_melted['Produto']
                    df2['SELL OUT MÊS'] = df_melted['QTD VENDA']
                    df2['ESTOQUE DISPONIVEL LOJA/CD'] = ''
                    df2['PENDENCIA TRANSITO/TRANSFERENCIA'] = ''
                    df2['PENDENCIA ENTREGA'] = ''
                    df2['ESTOQUE CD REDE'] = ''
                    
                    df_raw2 = pd.read_excel(xls, sheet_name='Estoque', header=None, skiprows=1)
                    # Extrair ano e mês da referência
                    ano = int(df_raw2.iloc[0, 1])
                    mes = int(df_raw2.iloc[1, 1])
                    data_ref = pd.to_datetime(f"{ano}-{mes:02d}-1").strftime('%m/%d/%Y')
                    # Localizar a linha de cabeçalho (onde está "Produto", "Cód Produto"...)
                    linha_cabecalho = df_raw2[df_raw2.iloc[:, 0].astype(str).str.strip() == "Produto"].index[0]
                    # Ler os dados reais com cabeçalho correto
                    df_estoque = pd.read_excel(xls, sheet_name='Estoque', header=linha_cabecalho, skiprows=1)
                    # Montar DataFrame final
                    df_estoque = df_estoque[
                            df_estoque['Produto'].notna() &
                            df_estoque['Cód Barra Venda'].notna()
                        ]

                    df_estoque2 = pd.DataFrame()
                    df_estoque2['EAN'] = normalizar_ean(df_estoque['Cód Barra Venda'].astype(str).str.zfill(13))
                    df_estoque2['MÊS'] = data_ref
                    df_estoque2['CANAL'] = 'REDE'
                    df_estoque2['CLIENTE'] = 'ARAUJO'
                    df_estoque2['GC'] = 'ANDRE RODRIGUES'
                    df_estoque2['UF'] = 'MG'
                    df_estoque2['FAMILIA'] = ''
                    df_estoque2['SKU_SERVIER'] = ''
                    df_estoque2['SELL OUT MÊS'] = ''
                    df_estoque2['ESTOQUE DISPONIVEL LOJA/CD'] = df_estoque['LOJAS']
                    df_estoque2['PENDENCIA TRANSITO/TRANSFERENCIA'] = ''
                    df_estoque2['PENDENCIA ENTREGA'] = ''
                    df_estoque2['ESTOQUE CD REDE'] = df_estoque['CD M. ANTONIO ARAUJO']
                    df2 = pd.merge(
                        df_estoque2,
                        df2,
                        on=['EAN', 'MÊS', 'CANAL', 'GC', 'UF', 'CLIENTE'],
                        how='outer',  # para manter todos os registros (vendas e/ou estoque)
                        suffixes=('_ESTOQUE', '_VENDA')
                    )

                    df2['FAMILIA'] = df2['FAMILIA_VENDA'].combine_first(df2['FAMILIA_ESTOQUE'])
                    df2['SKU_SERVIER'] = df2['SKU_SERVIER_VENDA'].combine_first(df2['SKU_SERVIER_ESTOQUE'])
                    df2['SELL OUT MÊS'] = df2['SELL OUT MÊS_VENDA'].combine_first(df2['SELL OUT MÊS_ESTOQUE'])
                    df2['ESTOQUE DISPONIVEL LOJA/CD'] = df2['ESTOQUE DISPONIVEL LOJA/CD_ESTOQUE'].combine_first(df2['ESTOQUE DISPONIVEL LOJA/CD_VENDA'])
                    df2['PENDENCIA TRANSITO/TRANSFERENCIA'] = df2['PENDENCIA TRANSITO/TRANSFERENCIA_ESTOQUE'].combine_first(df2['PENDENCIA TRANSITO/TRANSFERENCIA_VENDA'])
                    df2['PENDENCIA ENTREGA'] = df2['PENDENCIA ENTREGA_ESTOQUE'].combine_first(df2['PENDENCIA ENTREGA_VENDA'])
                    df2['ESTOQUE CD REDE'] = df2['ESTOQUE CD REDE_ESTOQUE'].combine_first(df2['ESTOQUE CD REDE_VENDA'])
                    df2.drop(columns=[col for col in df2.columns if col.endswith('_ESTOQUE') or col.endswith('_VENDA')], inplace=True)
                    # Caminho do arquivo de saída
                    output_file = os.path.join(output_path, 'Tabela_historico_STK.xlsx')

                    # Escrever no arquivo de forma segura
                    if os.path.exists(output_file):
                        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                            sheet_name = list(writer.sheets.keys())[0]
                            start_row = writer.sheets[sheet_name].max_row
                            df2.to_excel(writer, index=False, header=False, startrow=start_row)
                    else:
                        df2.to_excel(output_file, index=False)
                    df2.to_sql('tb_historico_stk', con=engine, if_exists='append', index=False, schema='stg')
                    gc.collect()
        except Exception as e:
            print(f"Erro ao ler {file}: {e}")
            return 0
