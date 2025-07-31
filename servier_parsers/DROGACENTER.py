import pandas as pd
import gc
import os
import zipfile
from openpyxl import load_workbook

def normalizar_ean(serie):
    return (
        serie.astype(str)                         # força string
            .astype(str)
            .str.extract(r'(\d+)')[0]   # pega só números
            .str[-13:]                  # mantém os 13 últimos dígitos
            .str.zfill(13)              # preenche zeros à esquerda se faltar
    )
def is_strict_openxml(filepath):
    try:
        with zipfile.ZipFile(filepath, 'r') as zipf:
            content_types = zipf.read('[Content_Types].xml').decode('utf-8')
            return 'Strict' in content_types
    except Exception:
        return False

def corrigir_strict_openxml(caminho_original, caminho_corrigido):
    try:
        wb = load_workbook(caminho_original)
        wb.save(caminho_corrigido)
        return True
    except Exception as e:
        print(f"Erro ao corrigir arquivo: {e}")
        return False
    
def tentar_ler_csv(filepath):
    encodings = ["utf-8", "latin1", "iso-8859-1", "cp1252"]
    separadores = [",", ";", "\t", "|"]
    for enc in encodings:
        for sep in separadores:
            try:
                df = pd.read_csv(filepath, encoding=enc, sep=sep)
                #print(f"CSV lido com encoding '{enc}' e separador '{sep}'")
                return df
            except Exception:
                continue
    raise ValueError(f"Falha ao ler CSV: {filepath} com encodings e separadores testados.")

def parse(file, caminho_completo, output_path, engine):
    if file.lower().endswith('.xlsx'):
        if is_strict_openxml(caminho_completo):
            print(f"Arquivo está no formato Strict Open XML: {file}")
            caminho_corrigido = caminho_completo.replace('.xlsx', '_corrigido.xlsx')
            if corrigir_strict_openxml(caminho_completo, caminho_corrigido):
                print(f"Corrigido e salvo como: {os.path.basename(caminho_corrigido)}")
                caminho_completo = caminho_corrigido
            else:
                print("Não foi possível corrigir automaticamente.")

        try:
            xls = pd.ExcelFile(caminho_completo, engine="openpyxl")
            for sheet in xls.sheet_names:
                print("Cliente - DROGACENTER")
                print(f"Arquivo - {file} | Aba - {sheet}")
                
                if 1 == 1:
                    # Ler a aba ignorando as 9 primeiras linhas
                    df_raw = pd.read_excel(xls, sheet_name=sheet,skiprows=3)
                    df_raw = df_raw[df_raw['Código EAN'].notna()]
                    df2 = pd.DataFrame()
                    df2['EAN'] = normalizar_ean(df_raw['Código EAN'].astype(str).str.zfill(13))
                    df2['MÊS'] = '02/01/2025'
                    df2['CANAL'] = 'DISTRIBUIDOR'
                    df2['CLIENTE'] = 'DROGACENTER'
                    df2['GC'] = 'ANDRE RODRIGUES'
                    df2['UF'] = 'SP'
                    df2['FAMILIA'] = df_raw['Produto'].astype(str).str.strip().str.split().str[0]
                    df2['SKU_SERVIER'] = df_raw['Produto']
                    df2['SELL OUT MÊS'] = df_raw['Vendas']
                    df2['ESTOQUE DISPONIVEL LOJA/CD'] = df_raw['Disponível']
                    df2['PENDENCIA TRANSITO/TRANSFERENCIA'] = df_raw['Andamento']
                    df2['PENDENCIA ENTREGA'] = ''
                    df2['ESTOQUE CD REDE'] = df_raw['Existente']
                    
                    
                    output_file = os.path.join(output_path, 'Tabela_historico_STK.xlsx')

                    # Escrever no arquivo de forma segura
                    if os.path.exists(output_file):
                        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                            sheet_name = list(writer.sheets.keys())[0]
                            start_row = writer.sheets[sheet_name].max_row
                            df2.to_excel(writer, index=False, header=False, startrow=start_row)
                    else:
                        df2.to_excel(output_file, index=False)
                    df2.to_sql('tb_historico_stk', con=engine, if_exists='append', index=False, schema='stg')
                    gc.collect()
        except Exception as e:
            print(f"Erro ao ler {file}: {e}")
            return 0
