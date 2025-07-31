import pandas as pd
import gc
import os
import zipfile
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore", message="Workbook contains no default style*")

def is_strict_openxml(filepath):
    try:
        with zipfile.ZipFile(filepath, 'r') as zipf:
            content_types = zipf.read('[Content_Types].xml').decode('utf-8')
            return 'Strict' in content_types
    except Exception:
        return False

def corrigir_strict_openxml(caminho_original, caminho_corrigido):
    try:
        wb = load_workbook(caminho_original)
        wb.save(caminho_corrigido)
        return True
    except Exception as e:
        print(f"Erro ao corrigir arquivo: {e}")
        return False
    
def tentar_ler_csv(filepath):
    encodings = ["utf-8", "latin1", "iso-8859-1", "cp1252"]
    separadores = [",", ";", "\t", "|"]
    for enc in encodings:
        for sep in separadores:
            try:
                df = pd.read_csv(filepath, encoding=enc, sep=sep)
                #print(f"CSV lido com encoding '{enc}' e separador '{sep}'")
                return df
            except Exception:
                continue
    raise ValueError(f"Falha ao ler CSV: {filepath} com encodings e separadores testados.")

def parse(file, caminho_completo, output_path, engine):
    if file.lower().endswith('.xlsx'):
            if is_strict_openxml(caminho_completo):
                print(f"Arquivo está no formato Strict Open XML: {file}")
                caminho_corrigido = caminho_completo.replace('.xlsx', '_corrigido.xlsx')
                if corrigir_strict_openxml(caminho_completo, caminho_corrigido):
                    print(f"Corrigido e salvo como: {os.path.basename(caminho_corrigido)}")
                    caminho_completo = caminho_corrigido
                else:
                    print("Não foi possível corrigir automaticamente.")

            # Leitura do Excel
            try:
                xls = pd.ExcelFile(caminho_completo, engine="openpyxl")
                print(xls.sheet_names)                
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    print("Cliente - ONCOPRODD")
                    print(f"Arquivo - {file} | Aba - {sheet}")
                    if file[:9] == "relatorio":
                        df2 = pd.DataFrame()
                        df2['INSTITUIÇÃO'] = df['Cliente PJ/Médico']
                        df2.insert(0,'DISTRIBUIDOR','ONCOPRODD')
                        df2.insert(1,'DATA',(pd.to_datetime(df['Data'], dayfirst=True)).dt.strftime('%m/%d/%Y'))
                        df2['CNPJ'] = df['CRM/CNPJ'].astype('str').str.zfill(14)
                        df2['EAN'] = df['Código Ean'].astype('str').str.zfill(13)
                        df2['CIDADE'] = df['Cidade'].str.upper()
                        df2['UF'] = df['UF CRM/CNPJ']
                        df2['QTD'] = df['Quantidade']
                        if os.path.exists(os.path.join(output_path, 'Tab_consolidado_demanda_Onco.xlsx')):
                            with pd.ExcelWriter(os.path.join(output_path, 'Tab_consolidado_demanda_Onco.xlsx'), mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                               df2.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                        else:
                            df2.to_excel(os.path.join(output_path, 'Tab_consolidado_demanda_Onco.xlsx'), index=False)
                        df2.to_sql('tb_demanda_onco', con=engine, if_exists='append', index=False, schema='stg')
                    elif file[:12] == "Mapa de esto":
                        df2 = pd.DataFrame()
                        df2['EAN'] = df['EAN'].astype('str').str.zfill(13)
                        df2.insert(0,'DISTRIBUIDOR','ONCOPRODD')
                        df2.insert(1,'DATA',(pd.to_datetime(df['Data Fechamento'], dayfirst=True)).dt.strftime('%m/%d/%Y'))
                        df2['TIPO'] = ''
                        df2['NOME DO CD'] = ''
                        df2['QTD ESTOQUE DISP'] = df['Estoque Livre'] + df['Estoque Livre Público'] + df['Estoque Consignado'] + df['Estoque Qualidade'] + df['Estoque Segregado GNV']
                        df2['QTD ESTOQUE TRANSITO'] = df['Pendência Trânsito']
                        df2['PEND. ENTREGA'] = df['Pendência Entrega'] + df['Pendência Entrega Público']
                        df2['CONS/EQUAL.'] = df['Transferência de Equalização']
                        df2['CIDADE'] = ''
                        df2['UF'] = df['CD'].str[:2]
                        df2['DATA VALIDADE'] = ''
                        if os.path.exists(os.path.join(output_path, 'Tab_consolidado_estoques_Onco.xlsx')):
                            with pd.ExcelWriter(os.path.join(output_path, 'Tab_consolidado_estoques_Onco.xlsx'), mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                               df2.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                        else:
                            df2.to_excel(os.path.join(output_path, 'Tab_consolidado_estoques_Onco.xlsx'), index=False)
                        df2.to_sql('tb_estoque_onco', con=engine, if_exists='append', index=False, schema='stg')
                    gc.collect()
            except Exception as e:
                print(f"Erro ao ler {file}: {e}")
                return 0